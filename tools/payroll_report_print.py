#!/usr/bin/env python3
"""Build a print-ready Payroll Report from the monthly workbook.

usage: payroll_report_print.py "Payroll Report_YYYY-MM_revN.xlsx" [out.html]

Reads the 'Payroll Report' sheet and emits A4-landscape HTML, one page per
business group, with the approval block on page 1.  Print from the browser
(Cmd-P -> Save as PDF), paper A4, orientation Landscape, margins Default.
"""
import sys, os, html, datetime
import openpyxl

SHEET = "Payroll Report"

# --- sheet geometry (1-based columns on the 'Payroll Report' sheet) ----------
GROUPS = [
    ("OPERATIONS",     list(range(3, 11)),  11),
    ("SUSTAINABILITY", list(range(12, 17)), 17),
    ("COMMERCIAL",     list(range(18, 21)), 21),
    ("EXPLORATION",    list(range(22, 24)), 24),
]
STANDALONE = [("BKK Office", 25), ("Legal", 26)]
GRAND_COL = 27

# Which groups land on which printed page.
PAGES = [
    ["OPERATIONS"],
    ["SUSTAINABILITY", "COMMERCIAL"],
    ["EXPLORATION", "_STANDALONE", "_GRAND"],
]

# Body rows: (kind, label, sheet_row)
#   kind 'sec'  section banner carrying that section's cost codes
#   kind 'hc'   headcount (integers)
#   kind 'amt'  money
#   kind 'tot'  money, emphasised subtotal
#   kind 'gt'   money, grand total rule
ROWS = [
    ("sec", "SENIOR STAFF", 9),
    ("hc",  "Headcount (จำนวนคน)", 10),
    ("amt", "Basic Salary", 11),
    ("amt", "Shift Allowance", 12),
    ("amt", "ERT Training", 13),
    ("amt", "Other Income", 14),
    ("tot", "Total — Senior Staff", 15),
    ("sec", "STAFF", 16),
    ("hc",  "Headcount (จำนวนคน)", 17),
    ("amt", "Basic Salary", 18),
    ("amt", "Shift Allowance", 19),
    ("amt", "Transportation", 20),
    ("amt", "Overtime", 21),
    ("amt", "ERT Training", 22),
    ("amt", "Other Income", 23),
    ("tot", "Total — Staff", 24),
    ("sec", "ANNUAL LEAVE (RESIGNED) & COMPENSATE", 25),
    ("amt", "Amount", 26),
    ("sec", "EMPLOYEES BONUS", 27),
    ("amt", "Amount", 28),
    ("sec", "PROVISION FOR SEVERANCE PAYMENTS", 29),
    ("amt", "Amount", 30),
    ("sec", "CONSULTANTS — TECHNICAL", 31),
    ("hc",  "Headcount (จำนวนคน)", 32),
    ("amt", "Amount", 33),
    ("sec", "CONTRACTORS — OTHER", 34),
    ("hc",  "Headcount (จำนวนคน)", 35),
    ("amt", "Amount", 36),
    ("sec", "CASUAL LABOUR", 37),
    ("hc",  "Headcount (จำนวนคน)", 38),
    ("amt", "Amount", 39),
    ("gt",  "GRAND TOTAL — PAYROLL EXPENSE", 40),
    ("gt",  "TOTAL HEADCOUNT (จำนวนคนทั้งหมด)", 41),
]

# Deduction block: single cost code per line, held in column B.
DED_ROWS = [
    ("amt", "Provident Fund", 44),
    ("amt", "Social Security", 45),
    ("amt", "Student Loan (general)", 46),
    ("amt", "Legal Execution Department", 47),
    ("amt", "Clearing Account — Employee W/Tax (PND 1)", 48),
    ("amt", "CL ACC EXP — Clearing Account W/Tax (PND 3)", 49),
    ("gt",  "GRAND TOTAL — DEDUCTION", 50),
]
NET_ROW = ("gt", "NET SALARY", 51)
PF_ROW = ("amt", "Provident Fund Employer Contribution", 54)


# --- helpers ----------------------------------------------------------------
def money(v):
    """Accounting format: thousands separated, 2dp, nil shown as a dash."""
    if v is None or v == "":
        return "–"
    if isinstance(v, str):
        return html.escape(v)
    if round(float(v), 2) == 0:
        return "–"
    return f"{float(v):,.2f}"


def count(v):
    if v is None or v == "" or (isinstance(v, (int, float)) and v == 0):
        return "–"
    if isinstance(v, str):
        return html.escape(v)
    return f"{int(round(float(v))):,}"


def code(v):
    return html.escape(str(v)) if v not in (None, "") else ""


def fmt_period(v):
    if isinstance(v, datetime.datetime):
        return v.strftime("%B %Y")
    return html.escape(str(v or ""))


def fmt_date(v):
    if isinstance(v, datetime.datetime):
        return v.strftime("%d %B %Y")
    return html.escape(str(v or ""))


def main():
    src = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else os.path.splitext(src)[0] + "_print.html"

    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb[SHEET]
    g = lambda r, c: ws.cell(r, c).value

    period = fmt_period(g(1, 23))
    repdate = fmt_date(g(2, 23))
    dept_name = {c: (g(8, c) or "").replace("\n", " ") for grp in GROUPS
                 for c in grp[1] + [grp[2]]}
    for name, c in STANDALONE:
        dept_name[c] = name
    dept_name[GRAND_COL] = "GRAND TOTAL"

    # columns for each printed page, in order, tagged with their group
    def page_columns(spec):
        cols, bands = [], []
        for key in spec:
            if key == "_STANDALONE":
                for name, c in STANDALONE:
                    bands.append((name, 1, False))
                    cols.append((c, True))
            elif key == "_GRAND":
                bands.append(("", 1, False))
                cols.append((GRAND_COL, True))
            else:
                title, dcols, tcol = next(x for x in GROUPS if x[0] == key)
                bands.append((title, len(dcols) + 1, True))
                cols += [(c, False) for c in dcols] + [(tcol, True)]
        return cols, bands

    P = []          # accumulated page HTML
    A = P.append

    # ---------- page 1 : summary + approval ----------
    tot_exp, tot_ded = g(5, 1), g(5, 8)
    tot_net, tot_hc = g(5, 14), g(5, 21)

    div_rows = ""
    for title, dcols, tcol in GROUPS:
        div_rows += (
            f"<tr><th>{html.escape(title)}</th>"
            f"<td>{money(g(40, tcol))}</td><td>{money(g(50, tcol))}</td>"
            f"<td>{money(g(51, tcol))}</td><td>{count(g(41, tcol))}</td></tr>")
    for name, c in STANDALONE:
        div_rows += (
            f"<tr><th>{html.escape(name)}</th>"
            f"<td>{money(g(40, c))}</td><td>{money(g(50, c))}</td>"
            f"<td>{money(g(51, c))}</td><td>{count(g(41, c))}</td></tr>")
    div_rows += (
        f"<tr class='gt'><th>GRAND TOTAL</th>"
        f"<td>{money(g(40, GRAND_COL))}</td><td>{money(g(50, GRAND_COL))}</td>"
        f"<td>{money(g(51, GRAND_COL))}</td><td>{count(g(41, GRAND_COL))}</td></tr>")

    # signature columns are wherever row 57 carries a heading — the block has
    # been re-merged between revisions, so never hard-code the columns.
    sig_cols = [c for c in range(1, 28) if g(57, c) not in (None, "")]
    sign = ""
    for c in sig_cols:
        sign += (
            "<div class='sig'>"
            f"<div class='sig-role'>{html.escape(str(g(57, c) or ''))}</div>"
            f"<div class='sig-unit'>{html.escape(str(g(58, c) or ''))}</div>"
            "<div class='sig-line'></div>"
            f"<div class='sig-name'>{html.escape(str(g(62, c) or '').replace('Name:', '').strip())}</div>"
            f"<div class='sig-pos'>{html.escape(str(g(63, c) or '').replace('Position:', '').strip())}</div>"
            "<div class='sig-date'>Date <span class='rule'></span></div>"
            "</div>")

    remarks = ""
    for r in range(69, ws.max_row + 1):
        if g(r, 1):
            remarks += (f"<tr><th>{html.escape(str(g(r,1)))}</th>"
                        f"<td>{html.escape(str(g(r,2) or ''))}</td></tr>")

    A(f"""<section class="page">
  <header class="masthead">
    <div class="brand"><h1>AKARA RESOURCES</h1>
      <p>Payroll Report — Employees &amp; Consultants</p></div>
    <div class="meta">
      <div><span>Payroll Period</span><strong>{period}</strong></div>
      <div><span>Report Date</span><strong>{repdate}</strong></div>
    </div>
  </header>

  <div class="kpis">
    <div class="kpi"><span>Total Payroll Expense</span><strong>{money(tot_exp)}</strong><em>THB</em></div>
    <div class="kpi"><span>Total Deduction</span><strong>{money(tot_ded)}</strong><em>THB</em></div>
    <div class="kpi accent"><span>Net Salary Payroll</span><strong>{money(tot_net)}</strong><em>THB</em></div>
    <div class="kpi"><span>Total Headcount</span><strong>{count(tot_hc)}</strong><em>persons</em></div>
  </div>

  <h2 class="sec-title">Summary by Business Group</h2>
  <table class="summary">
    <thead><tr><th>Business Group</th><th>Payroll Expense</th>
      <th>Deduction</th><th>Net Salary</th><th>Headcount</th></tr></thead>
    <tbody>{div_rows}</tbody>
  </table>

  <h2 class="sec-title">Approval</h2>
  <div class="signs">{sign}</div>

  <h2 class="sec-title">Remarks</h2>
  <table class="remarks"><tbody>{remarks}</tbody></table>
  <footer class="foot"><span>Payroll Report — {period}</span><span>Page 1 of {len(PAGES)+1}</span></footer>
</section>""")

    # ---------- detail pages ----------
    for pno, spec in enumerate(PAGES, start=2):
        cols, bands = page_columns(spec)
        ncol = len(cols)

        band_html = "<tr><th class='item' rowspan='2'>Item</th>"
        for title, span, _ in bands:
            band_html += (f"<th class='band' colspan='{span}'>{html.escape(title)}</th>"
                          if title else f"<th class='band empty' colspan='{span}'></th>")
        band_html += "</tr><tr>"
        for c, is_tot in cols:
            cls = "dept total" if is_tot else "dept"
            band_html += f"<th class='{cls}'>{html.escape(dept_name[c])}</th>"
        band_html += "</tr>"

        body = ""
        for kind, label, r in ROWS:
            if kind == "sec":
                body += (f"<tr class='sec'><th>{html.escape(label)}</th>"
                         + "".join(f"<td class='code'>{code(g(r,c))}</td>" for c, _ in cols)
                         + "</tr>")
            else:
                cls = {"hc": "hc", "amt": "amt", "tot": "amt sub", "gt": "amt gt"}[kind]
                f = count if kind == "hc" or "HEADCOUNT" in label.upper() else money
                body += (f"<tr class='{cls}'><th>{html.escape(label)}</th>"
                         + "".join(
                             f"<td{' class=totcol' if is_tot else ''}>{f(g(r,c))}</td>"
                             for c, is_tot in cols)
                         + "</tr>")

        # deduction + net + provident fund
        ded = f"<tr class='sec'><th>DEDUCTION — STAFF EXPENSES</th><td class='code' colspan='{ncol}'></td></tr>"
        for kind, label, r in DED_ROWS:
            cls = "amt gt" if kind == "gt" else "amt"
            cc = code(g(r, 2))
            lbl = (f"{html.escape(label)}<span class='cc'>{cc}</span>" if cc
                   else html.escape(label))
            ded += (f"<tr class='{cls}'><th>{lbl}</th>"
                    + "".join(f"<td{' class=totcol' if is_tot else ''}>{money(g(r,c))}</td>"
                              for c, is_tot in cols) + "</tr>")
        k, label, r = NET_ROW
        ded += (f"<tr class='amt net'><th>{html.escape(label)}</th>"
                + "".join(f"<td{' class=totcol' if is_tot else ''}>{money(g(r,c))}</td>"
                          for c, is_tot in cols) + "</tr>")

        k, label, r = PF_ROW
        pf = (f"<tr class='sec'><th>PROVIDENT FUND : K MASTER POOL FUND</th>"
              f"<td class='code' colspan='{ncol}'></td></tr>"
              f"<tr class='amt'><th>{html.escape(label)}"
              f"<span class='cc'>{code(g(r,2))}</span></th>"
              + "".join(f"<td{' class=totcol' if is_tot else ''}>{money(g(r,c))}</td>"
                        for c, is_tot in cols) + "</tr>")

        A(f"""<section class="page">
  <header class="masthead compact">
    <div class="brand"><h1>AKARA RESOURCES</h1>
      <p>Payroll Report — {" &amp; ".join(html.escape(t) for t,_,_ in bands if t)}</p></div>
    <div class="meta"><div><span>Payroll Period</span><strong>{period}</strong></div></div>
  </header>
  <div class="tablewrap">
  <table class="grid">
    <thead>{band_html}</thead>
    <tbody>{body}{ded}{pf}</tbody>
  </table>
  </div>
  <footer class="foot"><span>Payroll Report — {period}</span><span>Page {pno} of {len(PAGES)+1}</span></footer>
</section>""")

    doc = TEMPLATE.replace("__PERIOD__", period).replace("__BODY__", "\n".join(P))
    with open(out, "w", encoding="utf-8") as fh:
        fh.write(doc)
    print(f"wrote {out}")


TEMPLATE = """<!doctype html>
<html lang="en"><head>
<meta charset="utf-8">
<title>Payroll Report — __PERIOD__</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Sarabun:wght@400;500;600;700&display=swap" rel="stylesheet">
<style>
  @page { size: A4 landscape; margin: 9mm 10mm; }
  * { box-sizing: border-box; }
  html { -webkit-print-color-adjust: exact; print-color-adjust: exact; }
  body {
    margin: 0; background: #eef0f3;
    font-family: "Sarabun", "Helvetica Neue", Arial, sans-serif;
    color: #14181f;
  }
  .page {
    width: 277mm; min-height: 190mm; margin: 6mm auto; padding: 0;
    background: #fff; box-shadow: 0 1px 6px rgba(0,0,0,.18);
    display: flex; flex-direction: column;
  }
  @media print {
    body { background: #fff; }
    .page { width: auto; min-height: 0; margin: 0; box-shadow: none;
            page-break-after: always; break-after: page; }
    .page:last-child { page-break-after: auto; break-after: auto; }
  }

  /* ---- masthead ---- */
  .masthead {
    display: flex; justify-content: space-between; align-items: flex-end;
    border-bottom: 2.2pt solid #14181f; padding-bottom: 2.5mm; margin-bottom: 3.5mm;
  }
  .brand h1 { margin: 0; font-size: 15pt; font-weight: 700; letter-spacing: .14em; }
  .brand p  { margin: .6mm 0 0; font-size: 8.5pt; color: #4a5260; font-weight: 500; }
  .meta { display: flex; gap: 9mm; text-align: right; }
  .meta span { display: block; font-size: 6.4pt; letter-spacing: .13em;
               text-transform: uppercase; color: #6b7280; }
  .meta strong { font-size: 9.5pt; font-weight: 600; }
  .compact { padding-bottom: 1.6mm; margin-bottom: 2.2mm; border-bottom-width: 1.6pt; }
  .compact .brand h1 { font-size: 10.5pt; }
  .compact .brand p { font-size: 7.4pt; }
  .compact .meta strong { font-size: 8.5pt; }

  /* ---- kpis ---- */
  .kpis { display: grid; grid-template-columns: repeat(4, 1fr); gap: 3mm; margin-bottom: 5mm; }
  .kpi { border: .8pt solid #c9ced8; border-top: 2.4pt solid #14181f;
         padding: 3mm 3.5mm 3.2mm; }
  .kpi.accent { border-top-color: #1c5d3f; background: #f2f8f4; }
  .kpi span { display: block; font-size: 6.6pt; letter-spacing: .12em;
              text-transform: uppercase; color: #5b6472; margin-bottom: 1.6mm; }
  .kpi strong { display: block; font-size: 15pt; font-weight: 700;
                font-variant-numeric: tabular-nums; letter-spacing: -.01em; }
  .kpi.accent strong { color: #144c32; }
  .kpi em { font-style: normal; font-size: 6.6pt; color: #6b7280; letter-spacing: .1em; }

  .sec-title { font-size: 7.6pt; letter-spacing: .16em; text-transform: uppercase;
               color: #14181f; margin: 0 0 2mm; padding-bottom: 1.2mm;
               border-bottom: .8pt solid #14181f; font-weight: 700; }

  /* ---- summary table ---- */
  table { border-collapse: collapse; width: 100%; }
  .summary { font-size: 8.5pt; margin-bottom: 5mm; }
  .summary th, .summary td { padding: 1.8mm 3mm; border-bottom: .5pt solid #dfe3ea; }
  .summary thead th { font-size: 6.8pt; letter-spacing: .1em; text-transform: uppercase;
      color: #4a5260; text-align: right; border-bottom: 1pt solid #9aa3b2; }
  .summary thead th:first-child { text-align: left; }
  .summary tbody th { text-align: left; font-weight: 500; }
  .summary td { text-align: right; font-variant-numeric: tabular-nums; }
  .summary tr.gt th, .summary tr.gt td {
      font-weight: 700; border-top: 1pt solid #14181f;
      border-bottom: 2.2pt double #14181f; background: #f5f6f8; }

  /* ---- signatures ---- */
  .signs { display: grid; grid-auto-columns: 1fr; grid-auto-flow: column;\n           gap: 7mm; margin-bottom: 5mm; }
  .sig { border: .8pt solid #c9ced8; padding: 3mm 3.5mm 3mm; }
  .sig-role { font-size: 7pt; letter-spacing: .14em; text-transform: uppercase;
              font-weight: 700; }
  .sig-unit { font-size: 7pt; color: #6b7280; margin-bottom: 13mm; }
  .sig-line { border-bottom: .8pt solid #14181f; margin-bottom: 1.8mm; }
  .sig-name { font-size: 8.5pt; font-weight: 600; }
  .sig-pos  { font-size: 7.2pt; color: #4a5260; margin-bottom: 2.5mm; }
  .sig-date { font-size: 7.2pt; color: #4a5260; display: flex;
              align-items: flex-end; gap: 1.5mm; }
  .sig-date .rule { flex: 1; border-bottom: .6pt dotted #8a92a0; height: 3mm; }

  .remarks { font-size: 7.4pt; }
  .remarks th { text-align: left; font-weight: 600; width: 32mm;
                padding: 1.2mm 3mm 1.2mm 0; vertical-align: top; }
  .remarks td { color: #4a5260; padding: 1.2mm 0; }

  /* ---- detail grid ---- */
  .tablewrap { flex: 1; }
  .grid { font-size: 5.6pt; table-layout: fixed; line-height: 1.28; }
  .grid th, .grid td { border: .35pt solid #d5dae2; padding: .26mm 1mm;
                       font-variant-numeric: tabular-nums; }
  .grid thead th { background: #14181f; color: #fff; font-weight: 600;
                   text-align: center; vertical-align: middle; line-height: 1.15; }
  .grid thead .band { font-size: 6.3pt; letter-spacing: .1em; text-transform: uppercase;
                      border-color: #3a4150; }
  .grid thead .band.empty { background: #14181f; }
  .grid thead .dept { font-size: 5.5pt; font-weight: 500; background: #2b323e;
                      border-color: #3a4150; }
  .grid thead .dept.total { background: #47505f; font-weight: 700; }
  .grid thead .item { width: 41mm; text-align: left; padding-left: 1.7mm;
                      font-size: 6.3pt; letter-spacing: .08em; text-transform: uppercase; }
  .grid tbody th { text-align: left; font-weight: 400; padding-left: 2.2mm;
                   background: #fbfcfd; hyphens: none; }
  .grid tr.gt th { line-height: 1.15; }
  .grid tbody td { text-align: right; }
  .grid tbody td.totcol { background: #f1f3f6; font-weight: 600; }

  .grid tr.sec th { background: #dfe3ea; font-weight: 700; font-size: 5.4pt;
                    letter-spacing: .03em; text-transform: uppercase;
                    padding-left: 1.1mm; padding-top: .55mm; padding-bottom: .55mm; }
  .grid tr.sec td.code { background: #eef1f5; color: #58616f; font-size: 4.9pt;
                         text-align: center; letter-spacing: .01em; }
  .grid tr.hc td, .grid tr.hc th { color: #3a4150; }
  .grid tr.sub th, .grid tr.sub td { font-weight: 700; background: #f4f6f8; }
  .grid tr.sub td.totcol { background: #e8ecf1; }
  .grid tr.gt th, .grid tr.gt td { font-weight: 700; background: #14181f; color: #fff;
                                   border-color: #2b323e; }
  .grid tr.gt td.totcol { background: #2b323e; }
  .grid tr.net th, .grid tr.net td { font-weight: 700; background: #144c32; color: #fff;
                                     border-color: #1c5d3f; }
  .grid tr.net td.totcol { background: #1c5d3f; }
  .cc { display: inline; font-size: 4.9pt; color: #7b8492; letter-spacing: .02em;
        margin-left: 1.2mm; white-space: nowrap; }

  .foot { margin-top: auto; padding-top: 1.6mm; border-top: .6pt solid #c9ced8;
          display: flex; justify-content: space-between;
          font-size: 6.4pt; color: #6b7280; letter-spacing: .08em; }
</style>
</head><body>
__BODY__
</body></html>
"""

if __name__ == "__main__":
    main()
